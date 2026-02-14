"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  SeGaLog — Modèle MILP complet (scipy.optimize.milp / HiGHS)              ║
║  Planification stratégique supply chain sur 3 périodes T+0, T+1, T+2       ║
╚══════════════════════════════════════════════════════════════════════════════╝

MODÈLE MATHÉMATIQUE :
─────────────────────────────
Ensembles :
  Z = {AFR, AMN, AMS, ASI, EUR, OCE}
  P = {P1, P2, P3}
  S = {Petite, Moyenne, Grande}
  T = {0, 1, 2}

Variables de décision :
  u_{i,s,t} ∈ ℕ  : usines achetées en zone i, taille s, période t (effet en t+1)
  v_{i,s,t} ∈ ℕ  : usines fermées en zone i, taille s, période t (effet immédiat)
  N_{i,s,t} ∈ ℕ  : usines opérationnelles en zone i, taille s, période t
  x_{i,p,t} ∈ ℝ₊ : production du produit p en zone i, période t (unités)
  y_{i,j,p,t} ∈ ℝ₊: flux produit p de zone i vers zone j, période t (unités)
  n_{i,j,p,t} ∈ ℕ : nombre de conteneurs produit p de i vers j, période t

Objectif : Min Z_coût + Z_CO₂

Contraintes :
  C1. Satisfaction demande : ∑_i y_{i,j,p,t}×(1-ρ_{i,j,t}) ≥ D_{j,p,t}
  C2. Bilan production    : x_{i,p,t} = ∑_j y_{i,j,p,t}
  C3. Capacité (CORRIGÉE) : ∑_p x_{i,p,t}×h_p ≤ ∑_s N_{i,s,t}×CAP_s×TRS_{i,t}
  C4. Évolution usines    : N_{i,s,0} = N_init - v_{i,s,0}
                             N_{i,s,t} = N_{i,s,t-1} + u_{i,s,t-1} - v_{i,s,t}  (t≥1)
  C5. Limitation cession  : v_{i,s,t} ≤ usines disponibles
  C6. Liaison conteneurs  : y_{i,j,p,t} ≤ n_{i,j,p,t} × Q_p
"""

import numpy as np
from scipy.optimize import milp, LinearConstraint, Bounds
import openpyxl
import warnings
import sys
warnings.filterwarnings('ignore')

# ════════════════════════════════════════════════════════════════════════════
# CONSTANTES
# ════════════════════════════════════════════════════════════════════════════
ZONES    = ['AFR', 'AMN', 'AMS', 'ASI', 'EUR', 'OCE']
PRODUITS = ['P1', 'P2', 'P3']
TAILLES  = ['Petite', 'Moyenne', 'Grande']
PERIODES = [0, 1, 2]
EPSILON  = {0: 0.05, 1: 0.10, 2: 0.16}

nZ, nP, nS, nT = len(ZONES), len(PRODUITS), len(TAILLES), len(PERIODES)
Z_i = {z: i for i, z in enumerate(ZONES)}
P_i = {p: i for i, p in enumerate(PRODUITS)}
S_i = {s: i for i, s in enumerate(TAILLES)}


# ════════════════════════════════════════════════════════════════════════════
# 1. EXTRACTION DES DONNÉES
# ════════════════════════════════════════════════════════════════════════════
def extract_data(fich_donnees, fich_decisions):
    """Extrait toutes les données depuis les fichiers Excel."""
    data = {}

    for t in PERIODES:
        wb = openpyxl.load_workbook(fich_donnees, data_only=True)
        ws = wb[f'Données T+{t}']

        # Charger toutes les cellules dans un dict (row, col) -> value
        cells = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for c in row:
                if c.value is not None:
                    cells[(c.row, c.column)] = c.value

        def find(label, exact=False):
            """Trouve la ligne contenant label."""
            for (r, c_), v in cells.items():
                if c_ == 1 and isinstance(v, str):
                    if exact and v.strip() == label:
                        return r
                    if not exact and label in v:
                        return r
            return None

        def read_row(row, cols):
            return [float(cells.get((row, c), 0)) for c in cols]

        def read_matrix(header_row):
            """Lit matrice 6×6 sous la ligne header."""
            m = {}
            for i, zi in enumerate(ZONES):
                for j, zj in enumerate(ZONES):
                    v = cells.get((header_row + 1 + i, 2 + j), 0)
                    m[(zi, zj)] = float(v) if v else 0.0
            return m

        def read_zone_col(start_row, col=2):
            """Lit une colonne de 6 valeurs (une par zone) à partir de start_row."""
            d = {}
            # Chercher les zones dans les 8 lignes suivantes
            for offset in range(8):
                r = start_row + offset
                cv = cells.get((r, 1), '')
                if isinstance(cv, str) and cv.strip() in ZONES:
                    d[cv.strip()] = float(cells.get((r, col), 0))
            return d

        d = {}

        # Produits
        r = find('NB MO')
        d['h_p'] = dict(zip(PRODUITS, read_row(r, [2,3,4])))

        r = find('Cout Fixe')
        d['CF_p'] = dict(zip(PRODUITS, read_row(r, [2,3,4])))

        r = find('Prix de vente')
        d['Prix_vente'] = dict(zip(PRODUITS, read_row(r, [2,3,4])))

        r = find('Emission CO2')
        vals = []
        for c in [2,3,4]:
            v = cells.get((r, c), 0)
            if isinstance(v, str):
                v = float(v.replace('Kg','').replace('kg','').strip())
            vals.append(float(v))
        d['E_prod_p'] = dict(zip(PRODUITS, vals))

        # MO par zone
        r = find('Taux horaire')
        if r is None:
            r = find('MO', exact=True)
        d['MO'] = read_zone_col(r)

        # Transport
        r = find('Coût Conteneur')
        d['CT'] = read_matrix(r)

        # Conteneurs
        r = find('Qté')
        d['Q'] = dict(zip(PRODUITS, read_row(r, [2,3,4])))

        # Douanes
        r = find('Taux Impostion')
        d['tau'] = read_matrix(r)

        # Demande PIC
        r = find('Répartition demande')
        d['D_pic'] = {}
        d['D_hat'] = {}
        for i, z in enumerate(ZONES):
            for j, p in enumerate(PRODUITS):
                v = float(cells.get((r+1+i, 2+j), 0))
                d['D_pic'][(z, p)] = v
                d['D_hat'][(z, p)] = v * (1 + EPSILON[t])

        # Usines
        r = find('Capacité (/période)')
        d['CAP'] = dict(zip(TAILLES, read_row(r, [2,3,4])))

        r = find('Coût de fonct')
        d['CF_usine'] = dict(zip(TAILLES, read_row(r, [2,3,4])))

        r = find("Coût d'acquisition")
        d['CI'] = dict(zip(TAILLES, read_row(r, [2,3,4])))

        r = find('Coût de cession')
        d['CV'] = dict(zip(TAILLES, read_row(r, [2,3,4])))

        r = find('Emission CO2 Part Fixe')
        d['E_usine'] = dict(zip(TAILLES, read_row(r, [2,3,4])))

        # TRS
        r = find('Zones', exact=True)
        d['TRS'] = read_zone_col(r)

        # Risques pertes
        r = find('De / Vers')
        d['rho'] = read_matrix(r)

        # Stockage
        r = find('Cout de stock')
        d['CS'] = {}
        for jp, p in enumerate(PRODUITS):
            for k, z in enumerate(ZONES):
                d['CS'][(z, p)] = float(cells.get((r, 2 + jp*6 + k), 2.5))

        # Émissions transport
        r = find('T CO2 / Conteneur')
        d['E_transp'] = read_matrix(r)

        # Mix énergétique
        r = find('Coef')
        d['Mix'] = read_zone_col(r)

        # Prix CO₂
        r = find('Prix CO2/Tonnes')
        d['Prix_CO2'] = float(cells.get((r, 2), 95.81))

        data[t] = d
        wb.close()

    # Capacités initiales
    wb = openpyxl.load_workbook(fich_decisions, data_only=True)
    ws = wb['T0']
    N_init = {}
    for i, z in enumerate(ZONES):
        for j, s in enumerate(TAILLES):
            v = ws.cell(row=5+i, column=2+j).value
            try:
                n = int(float(v)) if v else 0
            except:
                n = 0
            if n > 0:
                N_init[(z, s)] = n
    data['N_init'] = N_init
    wb.close()

    return data


# ════════════════════════════════════════════════════════════════════════════
# 2. MODÈLE MILP
# ════════════════════════════════════════════════════════════════════════════
def build_and_solve(data, demand_mode='mixed', verbose=True):
    """
    Construit et résout le MILP.

    demand_mode:
      'nominal'    → utilise D_pic pour toutes les périodes
      'pessimistic'→ utilise D_hat = D_pic*(1+ε) pour toutes les périodes
      'mixed'      → D_pic pour T+0 (équilibré), D_hat pour T+1/T+2
    """

    # Indexation des variables
    # Par période t : u(nZ*nS) v(nZ*nS) N(nZ*nS) x(nZ*nP) y(nZ*nZ*nP) n(nZ*nZ*nP)
    nu = nZ*nS          # 18
    nv = nZ*nS          # 18
    nN = nZ*nS          # 18
    nx = nZ*nP          # 18
    ny = nZ*nZ*nP       # 108
    nn = nZ*nZ*nP       # 108
    block = nu+nv+nN+nx+ny+nn  # 288
    ntot = block * nT   # 864

    o_u = 0
    o_v = nu
    o_N = o_v + nv
    o_x = o_N + nN
    o_y = o_x + nx
    o_n = o_y + ny

    def iu(i,s,t): return t*block + o_u + i*nS + s
    def iv(i,s,t): return t*block + o_v + i*nS + s
    def iN(i,s,t): return t*block + o_N + i*nS + s
    def ix(i,p,t): return t*block + o_x + i*nP + p
    def iy(i,j,p,t): return t*block + o_y + (i*nZ+j)*nP + p
    def in_(i,j,p,t): return t*block + o_n + (i*nZ+j)*nP + p

    # Intégrité
    integrality = np.zeros(ntot, dtype=int)
    for t in range(nT):
        for i in range(nZ):
            for s in range(nS):
                integrality[iu(i,s,t)] = 1
                integrality[iv(i,s,t)] = 1
                integrality[iN(i,s,t)] = 1
            for j in range(nZ):
                for p in range(nP):
                    integrality[in_(i,j,p,t)] = 1

    # Bornes
    lb = np.zeros(ntot)
    ub = np.full(ntot, np.inf)
    for t in range(nT):
        for i in range(nZ):
            for s in range(nS):
                ub[iu(i,s,t)] = 3
                ub[iv(i,s,t)] = 3
                ub[iN(i,s,t)] = 5
        # Pas d'achat à T+2 (inutile, pas de T+3)
        if t == 2:
            for i in range(nZ):
                for s in range(nS):
                    ub[iu(i,s,t)] = 0

    # Fonction objectif
    c = np.zeros(ntot)
    for t in range(nT):
        pa = data[t]
        CF_p = pa['CF_p']; h_p = pa['h_p']; MO = pa['MO']
        CT = pa['CT']; tau = pa['tau']; Q = pa['Q']
        CF_u = pa['CF_usine']; CI = pa['CI']; CV = pa['CV']
        Eu = pa['E_usine']; Ep = pa['E_prod_p']; Et = pa['E_transp']
        Mix = pa['Mix']; PCO2 = pa['Prix_CO2']

        for i, zi in enumerate(ZONES):
            for p, pp in enumerate(PRODUITS):
                # Coût de production unitaire + coût CO₂ production
                cprod = CF_p[pp] + h_p[pp] * MO[zi]
                cco2  = PCO2 * Mix[zi] * Ep[pp] / 1000.0
                c[ix(i,p,t)] = cprod + cco2

            for s, ss in enumerate(TAILLES):
                # Coût usine : fonctionnement (k€→€) + CO₂ fixe usine
                c[iN(i,s,t)] = CF_u[ss]*1000 + PCO2 * Mix[zi] * Eu[ss]
                # Investissement
                c[iu(i,s,t)] = CI[ss]*1000
                # Cession (revenu = coût négatif)
                c[iv(i,s,t)] = -CV[ss]*1000

            for j, zj in enumerate(ZONES):
                for p, pp in enumerate(PRODUITS):
                    # Coût transport par conteneur + CO₂ transport
                    c[in_(i,j,p,t)] = CT[(zi,zj)] + PCO2 * Et[(zi,zj)]

                    # Douanes par unité transportée
                    if i != j and tau[(zi,zj)] > 0:
                        cprod_unit = CF_p[pp] + h_p[pp] * MO[zi]
                        ctransp_unit = CT[(zi,zj)] / Q[pp] if Q[pp] > 0 else 0
                        c[iy(i,j,p,t)] += (cprod_unit + ctransp_unit) * 1.1 * tau[(zi,zj)]

    # Contraintes
    eq_rows, eq_b = [], []
    ub_rows, ub_b = [], []

    def add_eq(coeffs, rhs):
        row = np.zeros(ntot)
        for idx, val in coeffs.items():
            row[idx] = val
        eq_rows.append(row); eq_b.append(rhs)

    def add_le(coeffs, rhs):
        row = np.zeros(ntot)
        for idx, val in coeffs.items():
            row[idx] = val
        ub_rows.append(row); ub_b.append(rhs)

    Ni = data['N_init']

    for t in range(nT):
        pa = data[t]

        # C4 : Évolution usines
        for i, zi in enumerate(ZONES):
            for s, ss in enumerate(TAILLES):
                if t == 0:
                    # N_{i,s,0} = N_init_{i,s} - v_{i,s,0}
                    add_eq({iN(i,s,0): 1, iv(i,s,0): 1}, Ni.get((zi,ss), 0))
                else:
                    # N_{i,s,t} = N_{i,s,t-1} + u_{i,s,t-1} - v_{i,s,t}
                    add_eq({iN(i,s,t): 1, iN(i,s,t-1): -1,
                            iu(i,s,t-1): -1, iv(i,s,t): 1}, 0)

        # C5 : Limitation cession
        for i, zi in enumerate(ZONES):
            for s, ss in enumerate(TAILLES):
                if t == 0:
                    add_le({iv(i,s,0): 1}, Ni.get((zi,ss), 0))
                else:
                    # v_{i,s,t} ≤ N_{i,s,t-1} + u_{i,s,t-1}
                    add_le({iv(i,s,t): 1, iN(i,s,t-1): -1, iu(i,s,t-1): -1}, 0)

        # C2 : Bilan production
        # x_{i,p,t} = ∑_j y_{i,j,p,t}
        for i in range(nZ):
            for p in range(nP):
                row = {ix(i,p,t): 1}
                for j in range(nZ):
                    row[iy(i,j,p,t)] = -1
                add_eq(row, 0)

        # C1 : Satisfaction demande
        for j, zj in enumerate(ZONES):
            for p, pp in enumerate(PRODUITS):
                if demand_mode == 'nominal':
                    D = pa['D_pic'][(zj,pp)]
                elif demand_mode == 'pessimistic':
                    D = pa['D_hat'][(zj,pp)]
                else:  # mixed
                    D = pa['D_pic'][(zj,pp)] if t == 0 else pa['D_hat'][(zj,pp)]
                if D <= 0:
                    continue
                # ∑_i y_{i,j,p,t} × (1 - ρ_{i,j}) ≥ D
                row = {}
                for i, zi in enumerate(ZONES):
                    rho = pa['rho'][(zi,zj)]  # ρ_{i,i}=0 dans les données
                    row[iy(i,j,p,t)] = -(1 - rho)
                add_le(row, -D)

        # C3 : Capacité production 
        for i, zi in enumerate(ZONES):
            row = {}
            for p, pp in enumerate(PRODUITS):
                row[ix(i,p,t)] = pa['h_p'][pp]
            for s, ss in enumerate(TAILLES):
                row[iN(i,s,t)] = -pa['CAP'][ss] * pa['TRS'][zi]
            add_le(row, 0)

        # C6 : Liaison conteneurs
        # y_{i,j,p,t} ≤ n_{i,j,p,t} × Q_p
        for i in range(nZ):
            for j in range(nZ):
                for p, pp in enumerate(PRODUITS):
                    add_le({iy(i,j,p,t): 1, in_(i,j,p,t): -pa['Q'][pp]}, 0)

    # Assemblage 
    A_eq = np.array(eq_rows) if eq_rows else None
    b_eq = np.array(eq_b)   if eq_rows else None
    A_ub = np.array(ub_rows) if ub_rows else None
    b_ub = np.array(ub_b)   if ub_rows else None

    constraints = []
    if A_eq is not None:
        constraints.append(LinearConstraint(A_eq, b_eq, b_eq))
    if A_ub is not None:
        constraints.append(LinearConstraint(A_ub, -np.inf, b_ub))

    if verbose:
        n_int = int(integrality.sum())
        n_cont = ntot - n_int
        print(f"  Variables      : {ntot} ({n_int} entières, {n_cont} continues)")
        print(f"  Contraintes éq : {len(eq_rows)}")
        print(f"  Contraintes ≤  : {len(ub_rows)}")
        print(f"  Mode demande   : {demand_mode}")

    # Résolution 
    if verbose:
        print("\n🔧 Lancement du solveur HiGHS...")

    result = milp(
        c=c,
        constraints=constraints,
        integrality=integrality,
        bounds=Bounds(lb, ub),
        options={"disp": verbose, "time_limit": 600, "mip_rel_gap": 0.005}
    )

    if not result.success:
        print(f"\n Échec : {result.message}")
        return None, None

    if verbose:
        print(f"\n Solution trouvée — Coût total : {result.fun:,.0f} €")

    # Extraction solution
    x = result.x
    sol = {'u':{}, 'v':{}, 'N':{}, 'x':{}, 'y':{}, 'n':{}, 'obj': result.fun}
    for t in range(nT):
        for i, zi in enumerate(ZONES):
            for s, ss in enumerate(TAILLES):
                sol['u'][(zi,ss,t)] = max(0, round(x[iu(i,s,t)]))
                sol['v'][(zi,ss,t)] = max(0, round(x[iv(i,s,t)]))
                sol['N'][(zi,ss,t)] = max(0, round(x[iN(i,s,t)]))
            for p, pp in enumerate(PRODUITS):
                sol['x'][(zi,pp,t)] = max(0, x[ix(i,p,t)])
            for j, zj in enumerate(ZONES):
                for p, pp in enumerate(PRODUITS):
                    sol['y'][(zi,zj,pp,t)] = max(0, x[iy(i,j,p,t)])
                    sol['n'][(zi,zj,pp,t)] = max(0, round(x[in_(i,j,p,t)]))

    return result, sol


# ════════════════════════════════════════════════════════════════════════════
# 3. AFFICHAGE
# ════════════════════════════════════════════════════════════════════════════
def afficher(sol, data, demand_mode='mixed'):
    if sol is None:
        print("Pas de solution."); return

    print("\n" + "═"*90)
    print(" RÉSULTATS PLANIFICATION STRATÉGIQUE SEGALOG")
    print("═"*90)

    total_rev = 0
    total_cost = sol['obj']

    for t in range(nT):
        pa = data[t]
        print(f"\n{'━'*90}")
        print(f" PÉRIODE T+{t}")
        print(f"{'━'*90}")

        # Usines achetées
        print("\n  Usines à ACHETER (opérationnelles en T+{})".format(min(t+1,2)))
        any_b = False
        for zi in ZONES:
            for ss in TAILLES:
                nb = sol['u'].get((zi,ss,t), 0)
                if nb > 0:
                    print(f"   {nb}× {ss} en {zi}  (invest: {nb*data[t]['CI'][ss]:,.0f} k€)")
                    any_b = True
        if not any_b:
            print("    Aucun achat")

        # Usines fermées
        print("\n  Usines à FERMER (effet immédiat)")
        any_s = False
        for zi in ZONES:
            for ss in TAILLES:
                nb = sol['v'].get((zi,ss,t), 0)
                if nb > 0:
                    print(f"   {nb}× {ss} en {zi}  (cession: {nb*data[t]['CV'][ss]:,.0f} k€)")
                    any_s = True
        if not any_s:
            print("    Aucune fermeture")

        # Parc usines
        print("\n  Parc d'usines opérationnelles")
        total_cap = 0
        for zi in ZONES:
            parts = []
            for ss in TAILLES:
                nb = sol['N'].get((zi,ss,t), 0)
                if nb > 0:
                    cap = nb * pa['CAP'][ss] * pa['TRS'][zi]
                    total_cap += cap
                    parts.append(f"{nb}×{ss} ({cap:,.0f} ut)")
            if parts:
                print(f"    {zi} : " + " | ".join(parts))

        # Demande en ut de capacité
        dem_cap = 0; dem_units = 0
        for zj in ZONES:
            for pp in PRODUITS:
                if demand_mode == 'nominal':
                    d = pa['D_pic'][(zj,pp)]
                elif demand_mode == 'pessimistic':
                    d = pa['D_hat'][(zj,pp)]
                else:
                    d = pa['D_pic'][(zj,pp)] if t == 0 else pa['D_hat'][(zj,pp)]
                dem_units += d
                dem_cap += d * pa['h_p'][pp]

        print(f"\n   Capacité totale : {total_cap:,.0f} ut")
        print(f"   Demande totale  : {dem_units:,.0f} produits = {dem_cap:,.0f} ut de capacité")
        marge = total_cap - dem_cap
        print(f"Marge capacitaire : {marge:,.0f} ut ({marge/total_cap*100:.1f}%)")

        # Production
        print(f"\n  Production par zone")
        for zi in ZONES:
            prods = []
            cap_used = 0
            for pp in PRODUITS:
                v = sol['x'].get((zi,pp,t), 0)
                if v > 0.5:
                    prods.append(f"{pp}={v:,.0f}")
                    cap_used += v * pa['h_p'][pp]
            if prods:
                print(f"    {zi} : " + " | ".join(prods) + f"  [capa utilisée: {cap_used:,.0f}]")

        # Flux
        print(f"\n Flux de distribution")
        for zi in ZONES:
            for zj in ZONES:
                for pp in PRODUITS:
                    q = sol['y'].get((zi,zj,pp,t), 0)
                    if q > 0.5:
                        nc = sol['n'].get((zi,zj,pp,t), 0)
                        print(f" {zi}→{zj} | {pp}: {q:,.0f} unités ({nc} conteneurs)")

        # Revenue
        rev = 0
        for zj in ZONES:
            for pp in PRODUITS:
                if demand_mode == 'nominal':
                    d = pa['D_pic'][(zj,pp)]
                elif demand_mode == 'pessimistic':
                    d = pa['D_hat'][(zj,pp)]
                else:
                    d = pa['D_pic'][(zj,pp)] if t == 0 else pa['D_hat'][(zj,pp)]
                rev += d * pa['Prix_vente'][pp]
        total_rev += rev
        print(f"\n    💰 CA période T+{t}: {rev:,.0f} €")

    # Bilan
    print(f"\n{'═'*90}")
    print(f" BILAN FINANCIER GLOBAL (3 périodes)")
    print(f"{'═'*90}")
    print(f"    Chiffre d'affaires : {total_rev:,.0f} €")
    print(f"    Coûts totaux       : {total_cost:,.0f} €")
    print(f"    BÉNÉFICE        : {total_rev - total_cost:,.0f} €")
    print(f"    Marge nette        : {(total_rev - total_cost)/total_rev*100:.1f}%")


def verifier_demande(sol, data, demand_mode='mixed'):
    print(f"\n{'═'*90}")
    print(" VÉRIFICATION SATISFACTION DEMANDE")
    print(f"{'═'*90}")
    ok = True
    for t in range(nT):
        pa = data[t]
        for zj in ZONES:
            for pp in PRODUITS:
                if demand_mode == 'nominal':
                    D = pa['D_pic'][(zj,pp)]
                elif demand_mode == 'pessimistic':
                    D = pa['D_hat'][(zj,pp)]
                else:
                    D = pa['D_pic'][(zj,pp)] if t == 0 else pa['D_hat'][(zj,pp)]
                if D <= 0:
                    continue
                recu = sum(
                    sol['y'].get((zi,zj,pp,t), 0) * (1 - pa['rho'][(zi,zj)])
                    for zi in ZONES
                )
                if recu < D - 1:
                    print(f" T+{t} {zj} {pp}: reçu={recu:,.0f} < demande={D:,.0f}")
                    ok = False
    if ok:
        print("  ✅ Toutes les demandes sont satisfaites !")
    return ok


def verifier_capacite(sol, data):
    print(f"\n{'═'*90}")
    print("  🔍  VÉRIFICATION CONTRAINTES DE CAPACITÉ")
    print(f"{'═'*90}")
    ok = True
    for t in range(nT):
        pa = data[t]
        for zi in ZONES:
            cap_used = sum(sol['x'].get((zi,pp,t),0)*pa['h_p'][pp] for pp in PRODUITS)
            cap_avail = sum(
                sol['N'].get((zi,ss,t),0) * pa['CAP'][ss] * pa['TRS'][zi]
                for ss in TAILLES
            )
            if cap_used > cap_avail + 1:
                print(f" T+{t} {zi}: utilisé={cap_used:,.0f} > dispo={cap_avail:,.0f}")
                ok = False
    if ok:
        print("  ✅ Toutes les capacités respectées !")
    return ok


# ════════════════════════════════════════════════════════════════════════════
# 4. EXPORT EXCEL — Feuille de décisions remplie
# ════════════════════════════════════════════════════════════════════════════
def export_decisions_excel(sol, data, src_path, dst_path, demand_mode='mixed'):
    """
    Copie le fichier de décisions original et le remplit avec la solution.
    """
    import shutil
    shutil.copy2(src_path, dst_path)

    wb = openpyxl.load_workbook(dst_path)

    sheet_names = {0: 'T0', 1: 'T1', 2: 'T2'}
    zone_rows = {z: 5+i for i, z in enumerate(ZONES)}  # AFR=5, AMN=6, ..., OCE=10
    prod_rows = {z: 17+i for i, z in enumerate(ZONES)}  # AFR=17, AMN=18, ..., OCE=22

    # Colonnes pour achat/fermeture : Petite=0, Moyenne=1, Grande=2
    achat_col_offset = {0: 6, 1: 7, 2: 8}      # F, G, H
    ferme_col_offset = {0: 10, 1: 11, 2: 12}    # J, K, L

    flux_col_base = {'P1': 2, 'P2': 8, 'P3': 14}

    for t in range(nT):
        ws = wb[sheet_names[t]]

        # Remplir achats et fermetures
        for zi in ZONES:
            row = zone_rows[zi]
            for s_idx, ss in enumerate(TAILLES):
                # Achat
                nb_achat = sol['u'].get((zi, ss, t), 0)
                ws.cell(row=row, column=achat_col_offset[s_idx], value=nb_achat)

                # Fermeture
                nb_ferme = sol['v'].get((zi, ss, t), 0)
                ws.cell(row=row, column=ferme_col_offset[s_idx], value=nb_ferme)

        # Remplir aussi les capacités initiales pour T0
        if t == 0:
            for zi in ZONES:
                row = zone_rows[zi]
                for s_idx, ss in enumerate(TAILLES):
                    n_init = data['N_init'].get((zi, ss), 0)
                    ws.cell(row=row, column=2+s_idx, value=n_init)

        # Remplir la matrice de répartition de production
        for i_idx, zi in enumerate(ZONES):
            row = prod_rows[zi]
            for pp in PRODUITS:
                base_col = flux_col_base[pp]
                for j_idx, zj in enumerate(ZONES):
                    col = base_col + j_idx
                    flux = sol['y'].get((zi, zj, pp, t), 0)
                    # Arrondir à l'entier (quantités de produits)
                    ws.cell(row=row, column=col, value=round(flux))

    wb.save(dst_path)
    print(f"\n Feuille de décisions remplie sauvegardée : {dst_path}")


# ════════════════════════════════════════════════════════════════════════════
# 5. MAIN
# ════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    fich_donnees   = 'SeGaLog_Donnee_v2.xlsx'
    fich_decisions = 'SeGaLoG_Feuille_de_decisions_v2.xlsx'

    print("═"*90)
    print("  SEGALOG — OPTIMISATION SUPPLY CHAIN MILP")
    print("═"*90)

    # ── Extraction ──
    print("\n Extraction des données...")
    data = extract_data(fich_donnees, fich_decisions)

    print("\n Capacités initiales :")
    for (z,s), n in sorted(data['N_init'].items()):
        print(f"   {z} — {s}: {n}")

    for t in PERIODES:
        d = data[t]
        cap_nom = sum(d['D_pic'][(z,p)]*d['h_p'][p] for z in ZONES for p in PRODUITS)
        cap_pes = sum(d['D_hat'][(z,p)]*d['h_p'][p] for z in ZONES for p in PRODUITS)
        print(f"\n T+{t}: Demande nominale = {cap_nom:,.0f} ut | pessimiste (+{EPSILON[t]*100:.0f}%) = {cap_pes:,.0f} ut")

    # Capacité initiale
    cap_init = sum(
        n * data[0]['CAP'][s] * data[0]['TRS'][z]
        for (z,s), n in data['N_init'].items()
    )
    print(f"\n Capacité initiale totale : {cap_init:,.0f} ut")

    # ── Résolution ──
    # Mode 'mixed'
    print("\n" + "─"*90)
    print("  RÉSOLUTION — Mode 'mixed' (nominal T+0, pessimiste T+1/T+2)")
    print("─"*90)
    result, sol = build_and_solve(data, demand_mode='mixed', verbose=True)

    if sol is not None:
        afficher(sol, data, demand_mode='mixed')
        verifier_demande(sol, data, demand_mode='mixed')
        verifier_capacite(sol, data)

        # ── Export Excel ──
        output_excel = 'SeGaLoG_Feuille_de_decisions_REMPLIE.xlsx'
        export_decisions_excel(sol, data, fich_decisions, output_excel, demand_mode='mixed')
    else:
        # Fallback : essayer en nominal pur
        print("\n Échec en mode mixed. Tentative en mode nominal...")
        result, sol = build_and_solve(data, demand_mode='nominal', verbose=True)
        if sol is not None:
            afficher(sol, data, demand_mode='nominal')
            verifier_demande(sol, data, demand_mode='nominal')
            verifier_capacite(sol, data)

            output_excel = 'SeGaLoG_Feuille_de_decisions_REMPLIE.xlsx'
            export_decisions_excel(sol, data, fich_decisions, output_excel, demand_mode='nominal')